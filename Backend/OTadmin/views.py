import pandas as pd
from django.shortcuts import render
from .models import OT
from datetime import datetime, timedelta, date
from .familias import mapa_familias
from .models import SegmentoCliente, normalizar_texto

import json
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseForbidden
from django.db.models import Sum, Count

from openpyxl import Workbook

from .kpi_service import (
    rango_default_30_dias,
    kpi_familias,
    kpi_segmentos,
    kpi_histograma_familias,
    kpi_cierres_temporales,
    kpi_globales,
)



def parsear_fecha(valor):
    """
    Devuelve una fecha válida o None.
    """

    if pd.isna(valor) or valor is None:
        return None

    # 1. Si es número de Excel
    if isinstance(valor, (int, float)):
        try:
            return datetime(1899, 12, 30) + timedelta(days=int(valor))
        except:
            pass

    # 2. Intentos comunes
    posibles_formatos = [
        "%d/%m/%Y",
        "%m/%d/%Y",
        "%d-%m-%Y",
        "%Y-%m-%d",
        "%d/%m/%y",
        "%m/%d/%y",
    ]

    for formato in posibles_formatos:
        try:
            return datetime.strptime(str(valor), formato)
        except:
            continue

    # 3. Parser automático final
    try:
        f = pd.to_datetime(valor, errors="coerce")
        if pd.isna(f):
            return None
        return f.to_pydatetime()
    except:
        return None

def subir_backlog(request):

    if request.method == "POST":
        archivo = request.FILES.get("archivo")

        if not archivo:
            return render(request, "subir_backlog.html", {"error": "No se seleccionó archivo."})

        try:
            # Leer primera pasada sin headers
            df_raw = pd.read_excel(archivo, header=None)

            # Detectar fila de encabezado
            header_row = df_raw[df_raw.apply(
                lambda row: row.astype(str).str.contains('SR_NO_INSTALACION').any(),
                axis=1
            )].index[0]

            df = pd.read_excel(archivo, header=header_row)
            df.columns = df.columns.str.strip().str.upper()

            # Filtrar solo VAS IMPLEMENTATION
            if "GRUPO" not in df.columns:
                return render(request, "subir_backlog.html", {"error": "El backlog no contiene la columna GRUPO."})

            df = df[df["GRUPO"] == "VAS IMPLEMENTATION"]

            agregadas = 0
            duplicadas = []

            for idx, row in df.iterrows():

                ot_num = row.get("SR_NO_INSTALACION")

                if pd.isna(ot_num):
                    continue

                # Evitar duplicados
                if OT.objects.filter(ot=ot_num).exists():
                    duplicadas.append(str(ot_num))
                    continue

                # =====================
                # 1) Cliente + normalización
                # =====================
                cliente = row.get("CLIENTE_NOMBRE", "")
                cliente_norm = normalizar_texto(cliente)

                # =====================
                # 2) Buscar segmento en tabla SegmentoCliente
                # =====================
                seg_obj = SegmentoCliente.objects.filter(cliente_normalizado=cliente_norm).first()
                segmento = seg_obj.segmento if seg_obj else ""

                # =====================
                # 3) Buscar familia (normalizado)
                # =====================
                producto = row.get("PRODUCTO", "")
                producto_norm = normalizar_texto(producto)
                familia = ""

                for clave, fam in mapa_familias.items():
                    if normalizar_texto(clave) == producto_norm:
                        familia = fam
                        break

                # =====================
                # 4) Parse fecha ingreso
                # =====================
                fecha = parsear_fecha(row.get("FECHA_LIBERACION"))
                if fecha is None:
                    fecha = datetime.today()

                # =====================
                # 5) Crear OT
                # =====================
                OT.objects.create(
                    ot=ot_num,
                    cliente=cliente,
                    cliente_normalizado=cliente_norm,
                    segmento=segmento,
                    producto=producto,
                    producto_hijo=row.get("SUB_PRODUCTO", ""),
                    familia=familia,
                    sol=row.get("ID_SOLUCION", ""),
                    enlace_id=row.get("NUMERO_DE_ENLACE"),
                    comercial=row.get("SR_USUARIO_CREADOR", ""),
                    fecha_ingreso=fecha,
                    fecha_estimada=(parsear_fecha(row.get("FECHA_LIBERACION")) is None)
                )

                agregadas += 1

            return render(request, "subir_backlog.html", {
                "ok": f"OT agregadas: {agregadas}",
                "duplicadas": duplicadas
            })

        except Exception as e:
            return render(request, "subir_backlog.html", {"error": str(e)})

    return render(request, "subir_backlog.html")


## PANTALLA VER SR ##

from django.core.paginator import Paginator
from django.db.models import Q
from datetime import datetime

def ver_sr(request):
    query = request.GET.get("q", "")
    estado = request.GET.get("estado", "")
    cliente = request.GET.get("cliente", "")
    producto = request.GET.get("producto", "")
    fecha_inicio = request.GET.get("fecha_inicio", "")
    fecha_fin = request.GET.get("fecha_fin", "")
    ing = request.GET.get("ing", "")

    ots = OT.objects.all().order_by("-fecha_ingreso")

    # Filtro general (barra de búsqueda)
    if query:
        ots = ots.filter(
            Q(ot__icontains=query) |
            Q(cliente__icontains=query) |
            Q(producto__icontains=query) |
            Q(sol__icontains=query) |
            Q(enlace_id__icontains=query)
        )

    # Filtro por estado
    if estado:
        ots = ots.filter(estado__iexact=estado)

    # Filtro por cliente
    if cliente:
        ots = ots.filter(cliente__icontains=cliente)

    # Filtro por producto
    if producto:
        ots = ots.filter(producto__icontains=producto)

    # Filtro por ingeniero responsable
    if ing:
        ots = ots.filter(ingeniero__icontains=ing)

    # Filtro por rango de fechas
    if fecha_inicio:
        ots = ots.filter(fecha_ingreso__gte=fecha_inicio)
    if fecha_fin:
        ots = ots.filter(fecha_ingreso__lte=fecha_fin)

    # Calcular días de cola
    for ot in ots:
        ot.dias = ot.dias_cola()
        ot.color = ot.color_cola()

    # PAGINACIÓN → 300 por página
    paginator = Paginator(ots, 300)
    page = request.GET.get("page")
    ots_paginated = paginator.get_page(page)

    return render(request, "ver_sr.html", {
        "ots": ots_paginated,
        "query": query,
        "estado": estado,
        "cliente": cliente,
        "producto": producto,
        "fecha_inicio": fecha_inicio,
        "fecha_fin": fecha_fin,
        "ing": ing,
        "paginator": paginator,
    })

## COMENTARIOS EN SR ##
from .models import OT, Comentario
from django.shortcuts import get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseForbidden


@login_required
def borrar_comentario(request, id):
    comentario = get_object_or_404(Comentario, id=id)

    # Solo el dueño puede borrar su comentario
    if comentario.usuario != request.user:
        return HttpResponseForbidden("No puedes borrar comentarios de otros usuarios.")

    ot_id = comentario.ot.id
    comentario.delete()

    return redirect("detalle_sr", ot_id=ot_id)

@login_required
def detalle_sr(request, ot_id):
    ot = get_object_or_404(OT, id=ot_id)

    # Crear comentario
    if request.method == "POST":
        texto = request.POST.get("comentario")
        if texto:
            Comentario.objects.create(
                ot=ot,
                usuario=request.user,
                texto=texto
            )
        return redirect("detalle_sr", ot_id=ot.id)

    # Obtener comentarios (solo 5 para inicio)
    comentarios = ot.comentarios.all()

    # Si tiene SOL, buscar otras OT con el mismo SOL
    ot_relacionadas = []
    if ot.sol:
        ot_relacionadas = OT.objects.filter(sol=ot.sol).exclude(id=ot.id)

    return render(request, "detalle_sr.html", {
        "ot": ot,
        "comentarios": comentarios,
        "ot_relacionadas": ot_relacionadas,
    })


## EDITAR SR ##
from datetime import date
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseForbidden
from django.utils.timezone import now

from accounts.models import CustomUser
from .models import OT, Comentario


@login_required
def editar_sr(request, ot_id):
    ot = get_object_or_404(OT, id=ot_id)

    if request.user.rol == "cliente":
        return HttpResponseForbidden("No tienes permisos para editar esta SR.")

    estados = OT.ESTADOS
    ingenieros = CustomUser.objects.filter(rol__in=["admin", "ing"])
    comentarios = ot.comentarios.order_by("-fecha")

    # Listas fijas (ajusta si luego vienen de BD)
    segmentos = ["MNC", "GOVERNMENT", "ENTERPRISE", "LARGE", "SMALL", "WHOLESALE"]
    familias = ["CLOUD", "CIBER", "FIJO", "SDWAN", "UCASS", "VAS"]

    if request.method == "POST":

        # Estado
        ot.estado = request.POST.get("estado")

        # Reglas por estado
        if ot.estado == "cerrado":
            ot.fecha_cierre = date.today()
        elif ot.estado in ["cancelado", "pospuesto"]:
            ot.fecha_cierre = None

        # Segmento y familia
        ot.segmento = request.POST.get("segmento")
        ot.familia = request.POST.get("familia")

        # Enlace
        ot.enlace_id = request.POST.get("enlace_id")

        # Ingeniero
        ing_id = request.POST.get("ing_encargado")
        ot.ing_encargado = (
            CustomUser.objects.get(id=ing_id) if ing_id else None
        )

        # Solo admin puede tocar fechas
        if request.user.rol == "admin":
            fecha_ing = request.POST.get("fecha_ingreso")
            fecha_cierre = request.POST.get("fecha_cierre")

            if fecha_ing:
                ot.fecha_ingreso = fecha_ing
            if fecha_cierre and ot.estado == "cerrado":
                ot.fecha_cierre = fecha_cierre

        ot.save()

        # Comentario
        comentario_texto = request.POST.get("nuevo_comentario")
        if comentario_texto:
            Comentario.objects.create(
                ot=ot,
                usuario=request.user,
                texto=comentario_texto,
                fecha=now(),
            )

        return redirect("detalle_sr", ot_id=ot.id)

    return render(request, "editar_sr.html", {
        "ot": ot,
        "estados": estados,
        "ingenieros": ingenieros,
        "comentarios": comentarios,
        "segmentos": segmentos,
        "familias": familias,
        "es_admin": request.user.rol == "admin",
    })


## AGREGAR USUARIO ##
import random
import string
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseForbidden
from django.shortcuts import render, redirect
from accounts.models import CustomUser


def generar_password(longitud=10):
    letras = string.ascii_letters
    numeros = string.digits
    simbolos = "!@#$%&*?"

    all_chars = letras + numeros + simbolos
    return "".join(random.choice(all_chars) for _ in range(longitud))


@login_required
def agregar_usuario(request):

    # Solo ADMIN puede entrar
    if request.user.rol != "admin":
        return HttpResponseForbidden("No tienes permisos para crear usuarios.")

    mensaje_ok = None
    mensaje_error = None

    if request.method == "POST":

        nombre = request.POST.get("nombre")
        apellido = request.POST.get("apellido")
        correo = request.POST.get("correo")
        rol = request.POST.get("rol")

        # Validar duplicado
        if CustomUser.objects.filter(email=correo).exists():
            mensaje_error = "Ya existe un usuario con ese correo."
        else:
            # Generar contraseña aleatoria
            password = generar_password()

            usuario = CustomUser.objects.create_user(
                username=correo,
                email=correo,
                nombre=nombre,
                apellido=apellido,
                rol=rol,
                password=password,
            )

            mensaje_ok = f"Usuario creado con éxito. Contraseña generada: {password}"

    return render(request, "agregar_usuario.html", {
        "mensaje_ok": mensaje_ok,
        "mensaje_error": mensaje_error,
    })

@login_required
def subir_segmento(request):
    from .models import SegmentoCliente, normalizar_texto

    if request.user.rol not in ["admin", "ing"]:
        return HttpResponseForbidden("No tienes permisos para esto.")

    if request.method == "POST":
        archivo = request.FILES.get("archivo")
        df = pd.read_excel(archivo)
        df.columns = df.columns.str.upper()

        # Tu Excel usa NOMBRE_CLIENTE y SEGMENTO
        for _, row in df.iterrows():
            cliente = row.get("NOMBRE_CLIENTE")
            segmento = row.get("SEGMENTO")

            if not cliente or not segmento:
                continue

            norm = normalizar_texto(cliente)

            SegmentoCliente.objects.update_or_create(
                cliente_normalizado=norm,
                defaults={
                    "cliente": cliente,
                    "segmento": segmento
                }
            )

        # Ahora sincronizamos TODAS las OT
        for ot in OT.objects.all():
            norm_ot = normalizar_texto(ot.cliente)
            seg = SegmentoCliente.objects.filter(cliente_normalizado=norm_ot).first()
            if seg:
                ot.segmento = seg.segmento
                ot.save()

        return render(request, "subir_segmento.html", {
            "ok": "Segmentos cargados y OTs sincronizadas correctamente."
        })

    return render(request, "subir_segmento.html")

@login_required
def ver_clientes(request):
    from .models import SegmentoCliente
    from django.core.paginator import Paginator

    if request.user.rol not in ["admin", "ing"]:
        return HttpResponseForbidden("No tienes permiso para ver esta página.")

    q = request.GET.get("q", "")

    clientes = SegmentoCliente.objects.all().order_by("cliente")

    # Buscador
    if q:
        clientes = clientes.filter(cliente__icontains=q)

    # PAGINACIÓN — 300 por página (igual que SR)
    paginator = Paginator(clientes, 300)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(request, "ver_clientes.html", {
        "clientes": page_obj,
        "q": q
    })

## VISTA KPI ##

@login_required
def graficas_view(request):
    # Solo Admin e Ing
    if request.user.rol not in ["admin", "ing"]:
        return HttpResponseForbidden("No tienes permisos para ver esta página.")

    fi_str = request.GET.get("fecha_inicio")
    ff_str = request.GET.get("fecha_fin")

    if fi_str and ff_str:
        try:
            fecha_inicio = datetime.strptime(fi_str, "%Y-%m-%d").date()
            fecha_fin = datetime.strptime(ff_str, "%Y-%m-%d").date()
        except ValueError:
            fecha_inicio, fecha_fin = rango_default_30_dias()
            fi_str = fecha_inicio.strftime("%Y-%m-%d")
            ff_str = fecha_fin.strftime("%Y-%m-%d")
    else:
        fecha_inicio, fecha_fin = rango_default_30_dias()
        fi_str = fecha_inicio.strftime("%Y-%m-%d")
        ff_str = fecha_fin.strftime("%Y-%m-%d")

    # Obtener datos de KPIs
    data_familias = kpi_familias(fecha_inicio, fecha_fin)
    data_segmentos = kpi_segmentos(fecha_inicio, fecha_fin)
    data_hist_familias = kpi_histograma_familias(fecha_inicio, fecha_fin)
    data_cierres = kpi_cierres_temporales(fecha_inicio, fecha_fin)
    data_globales = kpi_globales(fecha_inicio, fecha_fin)
    data_seg_comercial = kpi_segmentos_comercial_detallado(fecha_inicio, fecha_fin)

    context = {
        "fecha_inicio": fi_str,
        "fecha_fin": ff_str,
        "familias_json": json.dumps(data_familias),
        "segmentos_json": json.dumps(data_segmentos),
        "hist_familias_json": json.dumps(data_hist_familias),
        "cierres_json": json.dumps(data_cierres),
        "seg_comercial_json": json.dumps(data_seg_comercial),
        "globales": data_globales,
    }

    return render(request, "graficas.html", context)


@login_required
def exportar_kpis_excel(request):
    if request.user.rol not in ["admin", "ing"]:
        return HttpResponseForbidden("No tienes permisos para esto.")

    fi_str = request.GET.get("fecha_inicio")
    ff_str = request.GET.get("fecha_fin")

    if fi_str and ff_str:
        try:
            fecha_inicio = datetime.strptime(fi_str, "%Y-%m-%d").date()
            fecha_fin = datetime.strptime(ff_str, "%Y-%m-%d").date()
        except ValueError:
            fecha_inicio, fecha_fin = rango_default_30_dias()
    else:
        fecha_inicio, fecha_fin = rango_default_30_dias()

    # Reusar servicios KPI
    data_familias = kpi_familias(fecha_inicio, fecha_fin)
    data_segmentos = kpi_segmentos(fecha_inicio, fecha_fin)
    data_hist_familias = kpi_histograma_familias(fecha_inicio, fecha_fin)
    data_cierres = kpi_cierres_temporales(fecha_inicio, fecha_fin)
    data_globales = kpi_globales(fecha_inicio, fecha_fin)

    wb = Workbook()
    wb.remove(wb.active)  # borrar hoja por defecto

    # 1) KPIs_Familias
    ws1 = wb.create_sheet("KPIs_Familias")
    ws1.append(["Familia", "OTs cerradas", "MTTI promedio", "% del total"])
    for row in data_familias:
        ws1.append([
            row["familia"],
            row["total"],
            row["mtti_promedio"],
            row["porcentaje"],
        ])

    # 2) KPIs_Segmentos
    ws2 = wb.create_sheet("KPIs_Segmentos")
    ws2.append(["Segmento", "OTs cerradas", "MTTI promedio", "% del total"])
    for row in data_segmentos:
        ws2.append([
            row["segmento"],
            row["total"],
            row["mtti_promedio"],
            row["porcentaje"],
        ])

    # 3) Distribucion_Familias
    ws3 = wb.create_sheet("Distribucion_Familias")
    ws3.append(["Familia", "0-8 días", "9-16 días", "17-20 días", ">20 días"])
    for row in data_hist_familias:
        ws3.append([
            row["familia"],
            row["rango_0_8"],
            row["rango_9_16"],
            row["rango_17_20"],
            row["rango_mas_20"],
        ])

    # 4) Cierres_Temporales
    ws4 = wb.create_sheet("Cierres_Temporales")
    ws4.append(["Periodo (YYYY-MM)", "OTs cerradas", "MTTI promedio"])
    for row in data_cierres:
        ws4.append([
            row["periodo"],
            row["total"],
            row["mtti_promedio"],
        ])

    # 5) Globales
    ws5 = wb.create_sheet("Globales")
    ws5.append(["Métrica", "Valor"])
    ws5.append(["Total OTs cerradas", data_globales["total_ot_cerradas"]])
    ws5.append(["MTTI promedio", data_globales["mtti_promedio"]])
    ws5.append(["Fecha inicio", fecha_inicio.strftime("%Y-%m-%d")])
    ws5.append(["Fecha fin", fecha_fin.strftime("%Y-%m-%d")])

    # 6) Datos_crudos (opcional, para validar)
    qs = qs_cerradas_en_rango(fecha_inicio, fecha_fin)
    ws6 = wb.create_sheet("Datos_crudos")
    ws6.append([
        "OT", "Cliente", "Segmento", "Familia",
        "Fecha ingreso", "Fecha cierre", "MTTI",
    ])
    for ot in qs:
        ws6.append([
            ot.ot,
            ot.cliente,
            ot.segmento,
            ot.familia,
            ot.fecha_ingreso.strftime("%Y-%m-%d") if ot.fecha_ingreso else "",
            ot.fecha_cierre.strftime("%Y-%m-%d") if ot.fecha_cierre else "",
            ot.mtti,
        ])

    # Respuesta HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"kpis_OT_{fecha_inicio}_{fecha_fin}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


# ====== RANGO DEFAULT DE 30 DÍAS (ÚLTIMOS 30) ======
def rango_default_30_dias():
    hoy = date.today()
    inicio = hoy - timedelta(days=30)
    return inicio, hoy


# ====== QS BASE PARA KPIs: SOLO OTs CERRADAS EN RANGO ======
def qs_cerradas_en_rango(fecha_inicio, fecha_fin):
    return OT.objects.filter(
        estado="cerrado",
        fecha_cierre__isnull=False,
        fecha_cierre__range=(fecha_inicio, fecha_fin),
    )

FAMILIAS_KPI = ["CLOUD", "CIBER", "FIJO", "SDWAN", "UCASS", "VAS"]
SEGMENTOS_KPI = ["SMALL", "LARGE", "ENTERPRISE", "GOVERNMENT", "WHOLESALE", "MNC"]


def kpi_familias(fecha_inicio, fecha_fin):
    qs = qs_cerradas_en_rango(fecha_inicio, fecha_fin).exclude(familia__isnull=True).exclude(familia="")
    data = {fam: {"total": 0, "suma_mtti": 0, "conteo_mtti": 0} for fam in FAMILIAS_KPI}

    for ot in qs:
        fam = (ot.familia or "").upper()
        if fam not in data:
            continue
        data[fam]["total"] += 1
        if ot.mtti is not None:
            data[fam]["suma_mtti"] += ot.mtti
            data[fam]["conteo_mtti"] += 1

    total_global = sum(v["total"] for v in data.values()) or 1

    resultado = []
    for fam in FAMILIAS_KPI:
        info = data[fam]
        if info["conteo_mtti"] > 0:
            mtti_prom = round(info["suma_mtti"] / info["conteo_mtti"], 1)
        else:
            mtti_prom = 0

        porcentaje = round(100 * info["total"] / total_global, 1) if info["total"] > 0 else 0

        resultado.append({
            "familia": fam,
            "total": info["total"],
            "mtti_promedio": mtti_prom,
            "porcentaje": porcentaje,
        })

    return resultado


def kpi_segmentos(fecha_inicio, fecha_fin):
    qs = qs_cerradas_en_rango(fecha_inicio, fecha_fin).exclude(segmento__isnull=True).exclude(segmento="")
    data = {seg: {"total": 0, "suma_mtti": 0, "conteo_mtti": 0} for seg in SEGMENTOS_KPI}

    for ot in qs:
        seg = (ot.segmento or "").upper()
        if seg not in data:
            continue
        data[seg]["total"] += 1
        if ot.mtti is not None:
            data[seg]["suma_mtti"] += ot.mtti
            data[seg]["conteo_mtti"] += 1

    total_global = sum(v["total"] for v in data.values()) or 1

    resultado = []
    for seg in SEGMENTOS_KPI:
        info = data[seg]
        if info["conteo_mtti"] > 0:
            mtti_prom = round(info["suma_mtti"] / info["conteo_mtti"], 1)
        else:
            mtti_prom = 0

        porcentaje = round(100 * info["total"] / total_global, 1) if info["total"] > 0 else 0

        resultado.append({
            "segmento": seg,
            "total": info["total"],
            "mtti_promedio": mtti_prom,
            "porcentaje": porcentaje,
        })

    return resultado


def kpi_histograma_familias(fecha_inicio, fecha_fin):
    """
    Devuelve por familia:
    - rangos: 0-8 / 9-16 / 17-20 / >20
    - y además bins/valores para Plotly
    """
    qs = qs_cerradas_en_rango(fecha_inicio, fecha_fin).exclude(familia__isnull=True).exclude(familia="")
    # Solo mtti definidos
    qs = qs.exclude(mtti__isnull=True)

    # Inicializar contadores
    data = {
        fam: {
            "rango_0_8": 0,
            "rango_9_16": 0,
            "rango_17_20": 0,
            "rango_mas_20": 0,
        } for fam in FAMILIAS_KPI
    }

    for ot in qs:
        fam = (ot.familia or "").upper()
        if fam not in data:
            continue

        d = ot.mtti
        if d is None:
            continue
        if d <= 8:
            data[fam]["rango_0_8"] += 1
        elif d <= 16:
            data[fam]["rango_9_16"] += 1
        elif d <= 20:
            data[fam]["rango_17_20"] += 1
        else:
            data[fam]["rango_mas_20"] += 1

    resultado = []
    for fam in FAMILIAS_KPI:
        info = data[fam]
        bins = ["0-8", "9-16", "17-20", ">20"]
        valores = [
            info["rango_0_8"],
            info["rango_9_16"],
            info["rango_17_20"],
            info["rango_mas_20"],
        ]

        resultado.append({
            "familia": fam,
            "rango_0_8": info["rango_0_8"],
            "rango_9_16": info["rango_9_16"],
            "rango_17_20": info["rango_17_20"],
            "rango_mas_20": info["rango_mas_20"],
            "bins": bins,
            "valores": valores,
        })

    return resultado


def kpi_cierres_temporales(fecha_inicio, fecha_fin):
    """
    Agrupa las OTs cerradas por mes (YYYY-MM) con total y MTTI promedio.
    """
    qs = qs_cerradas_en_rango(fecha_inicio, fecha_fin).exclude(mtti__isnull=True)

    buckets = {}
    for ot in qs:
        periodo = ot.fecha_cierre.strftime("%Y-%m")
        if periodo not in buckets:
            buckets[periodo] = {"total": 0, "suma_mtti": 0, "conteo_mtti": 0}
        buckets[periodo]["total"] += 1
        buckets[periodo]["suma_mtti"] += ot.mtti
        buckets[periodo]["conteo_mtti"] += 1

    resultado = []
    for periodo in sorted(buckets.keys()):
        info = buckets[periodo]
        if info["conteo_mtti"] > 0:
            mtti_prom = round(info["suma_mtti"] / info["conteo_mtti"], 1)
        else:
            mtti_prom = 0

        resultado.append({
            "periodo": periodo,
            "total": info["total"],
            "mtti_promedio": mtti_prom,
        })

    return resultado


def kpi_globales(fecha_inicio, fecha_fin):
    qs = qs_cerradas_en_rango(fecha_inicio, fecha_fin).exclude(mtti__isnull=True)
    total = qs.count()
    if total == 0:
        mtti_prom = 0
    else:
        suma = qs.aggregate(s=Sum("mtti"))["s"] or 0
        mtti_prom = round(suma / total, 1)

    return {
        "total_ot_cerradas": total,
        "mtti_promedio": mtti_prom,
    }


def kpi_segmentos_comercial_detallado(fecha_inicio, fecha_fin):
    """
    Para la gráfica tipo PDF:
    Para cada segmento (MNC, GOV, ENTERPRISE, LARGE, SMALL, WHOLESALE)
    arma:
        - familias: ["CIBER","CLOUD",...]
        - sr: [conteos]
        - mtti: [promedio mtti]
    """
    qs = qs_cerradas_en_rango(fecha_inicio, fecha_fin).exclude(mtti__isnull=True)

    resultado = []

    for seg in SEGMENTOS_KPI:
        fam_labels = []
        sr_counts = []
        mtti_vals = []

        for fam in FAMILIAS_KPI:
            ots_sf = [
                ot for ot in qs
                if (ot.segmento or "").upper() == seg
                and (ot.familia or "").upper() == fam
            ]
            total = len(ots_sf)
            if total == 0:
                continue

            suma_mtti = sum(ot.mtti for ot in ots_sf if ot.mtti is not None)
            if total > 0:
                mtti_prom = round(suma_mtti / total, 1)
            else:
                mtti_prom = 0

            fam_labels.append(fam)
            sr_counts.append(total)
            mtti_vals.append(mtti_prom)

        if fam_labels:
            resultado.append({
                "segmento": seg,
                "familias": fam_labels,
                "sr": sr_counts,
                "mtti": mtti_vals,
            })

    return resultado
